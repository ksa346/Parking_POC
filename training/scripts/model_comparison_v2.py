"""
Standalone Model Comparison Script
Evaluates RF-DETR model performance on videos with ground truth car counts.
NO API CALLS - Direct model loading and inference.

Usage:
    python model_comparison_v2.py --video parking.mp4 --model base --ground-truth 25
    python model_comparison_v2.py --video parking.mp4 --model checkpoint.pth --ground-truth 30

Output:
    Model_Comparison.xlsx with metrics in Sheet1
"""

import argparse
import sys
import os
import importlib
import tempfile
from pathlib import Path
from typing import Dict, List, Tuple
import logging
import time

import cv2
import numpy as np
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from scipy.spatial.distance import cdist

try:
    import torch
    TORCH_AVAILABLE = True
except ImportError:
    TORCH_AVAILABLE = False
    print("WARNING: PyTorch not available")

try:
    # Newer rfdetr builds expose model variants via rfdetr.variants.
    from rfdetr.variants import RFDETRBase, RFDETRSmall, RFDETRMedium, RFDETRLarge, RFDETRNano
    RFDETR_AVAILABLE = True
    RFDETR_IMPORT_ERROR = None
except Exception:
    try:
        # Fallback for older packaging layouts.
        from rfdetr import RFDETRBase, RFDETRSmall, RFDETRMedium, RFDETRLarge, RFDETRNano
        RFDETR_AVAILABLE = True
        RFDETR_IMPORT_ERROR = None
    except Exception:
        RFDETR_AVAILABLE = False
        RFDETR_IMPORT_ERROR = str(sys.exc_info()[1])

logging.basicConfig(level=logging.INFO, format='%(levelname)s: %(message)s')
logger = logging.getLogger(__name__)


class SimpleTracker:
    """Simple centroid tracker for vehicle counting."""
    
    def __init__(self, max_distance=50, max_age=30):
        self.next_id = 0
        self.tracks = {}  # {id: {'centroid': (x,y), 'prev_centroid': (x,y), 'age': int, 'counted': bool}}
        self.max_distance = max_distance
        self.max_age = max_age
    
    def update(self, boxes):
        """Update with new boxes. Returns track IDs."""
        if not boxes:
            # Age out all tracks
            self.tracks = {k: v for k, v in self.tracks.items() if v['age'] < self.max_age}
            return {}
        
        # Calculate centroids
        centroids = []
        for (x1, y1, x2, y2) in boxes:
            cx = (x1 + x2) / 2.0
            cy = (y1 + y2) / 2.0
            centroids.append([cx, cy])
        centroids = np.array(centroids)
        
        if len(self.tracks) == 0:
            # Initialize tracks
            for centroid in centroids:
                self.tracks[self.next_id] = {
                    'centroid': centroid,
                    'prev_centroid': centroid.copy(),
                    'age': 0,
                    'counted': False,
                }
                self.next_id += 1
        else:
            # Match centroids to existing tracks
            track_ids = list(self.tracks.keys())
            track_centroids = np.array([self.tracks[tid]['centroid'] for tid in track_ids])
            
            # Distance matrix
            D = cdist(track_centroids, centroids)
            
            matched_track = set()
            matched_det = set()
            
            for t_idx, t_id in enumerate(track_ids):
                if t_idx >= D.shape[0]:
                    continue
                min_dist_idx = np.argmin(D[t_idx])
                min_dist = D[t_idx, min_dist_idx]
                
                if min_dist < self.max_distance and min_dist_idx not in matched_det:
                    self.tracks[t_id]['prev_centroid'] = self.tracks[t_id]['centroid']
                    self.tracks[t_id]['centroid'] = centroids[min_dist_idx]
                    self.tracks[t_id]['age'] = 0
                    matched_track.add(t_id)
                    matched_det.add(min_dist_idx)
            
            # Age out unmatched tracks
            for t_id in track_ids:
                if t_id not in matched_track:
                    self.tracks[t_id]['age'] += 1
                    if self.tracks[t_id]['age'] > self.max_age:
                        del self.tracks[t_id]
            
            # Create new tracks for unmatched detections
            for d_idx, centroid in enumerate(centroids):
                if d_idx not in matched_det:
                    self.tracks[self.next_id] = {
                        'centroid': centroid,
                        'prev_centroid': centroid.copy(),
                        'age': 0,
                        'counted': False,
                    }
                    self.next_id += 1
        
        return self.tracks
    
    def count(self):
        """Count active tracks."""
        return len(self.tracks)

    @staticmethod
    def _point_side(point, p1, p2):
        """Signed side of point relative to line p1->p2."""
        x, y = point
        x1, y1 = p1
        x2, y2 = p2
        return (x - x1) * (y2 - y1) - (y - y1) * (x2 - x1)

    def count_tripwire_crossings(self, p1, p2):
        """Count tracks that crossed the tripwire since last update (count each track once)."""
        crossings = 0
        for track in self.tracks.values():
            if track.get('counted', False):
                continue

            prev_pt = track.get('prev_centroid')
            curr_pt = track.get('centroid')
            if prev_pt is None or curr_pt is None:
                continue

            prev_side = self._point_side(prev_pt, p1, p2)
            curr_side = self._point_side(curr_pt, p1, p2)

            # Strict sign change indicates crossing between frames.
            if prev_side * curr_side < 0:
                track['counted'] = True
                crossings += 1

        return crossings


def parse_trip_wire(trip_wire_spec: str):
    """Parse trip-wire text like {x1=700;y1=300;x2=1600;y2=300}."""
    if not trip_wire_spec:
        return None

    text = trip_wire_spec.strip()
    if text.startswith('{') and text.endswith('}'):
        text = text[1:-1]

    parts = [p.strip() for p in text.split(';') if p.strip()]
    values = {}
    for part in parts:
        if '=' not in part:
            raise ValueError(f"Invalid trip-wire segment '{part}'. Expected key=value")
        key, value = part.split('=', 1)
        values[key.strip().lower()] = int(float(value.strip()))

    required = ('x1', 'y1', 'x2', 'y2')
    missing = [k for k in required if k not in values]
    if missing:
        raise ValueError(f"Missing trip-wire keys: {', '.join(missing)}")

    return (values['x1'], values['y1']), (values['x2'], values['y2'])


def load_model(model_spec: str):
    """Load RF-DETR model. Supports 'base', 'small', etc. or checkpoint paths."""
    if not RFDETR_AVAILABLE:
        conda_env = os.environ.get("CONDA_DEFAULT_ENV", "(not set)")
        raise RuntimeError(
            "rfdetr import failed in the current interpreter.\n"
            f"Python executable: {sys.executable}\n"
            f"Python version: {sys.version.split()[0]}\n"
            f"CONDA_DEFAULT_ENV: {conda_env}\n"
            f"Import error: {RFDETR_IMPORT_ERROR}\n"
            "Fix: install into THIS interpreter: python -m pip install rfdetr\n"
            "Or run with conda explicitly: conda run -n <your_env> python model_comparison_v2.py ..."
        )
    
    if not TORCH_AVAILABLE:
        raise RuntimeError("PyTorch not installed. Run: pip install torch")
    
    logger.info(f"Loading model: {model_spec}")

    def _get_rfdetr_class(name: str):
        try:
            module = importlib.import_module("rfdetr")
            cls = getattr(module, name, None)
            if cls is not None:
                return cls
        except Exception:
            pass
        for module_name in ("rfdetr.variants", "rfdetr.detr"):
            try:
                submod = importlib.import_module(module_name)
                cls = getattr(submod, name, None)
                if cls is not None:
                    return cls
            except Exception:
                continue
        return None
    
    # Check if checkpoint file
    if os.path.isfile(model_spec) and model_spec.endswith(('.pth', '.pt', '.ckpt')):
        logger.info(f"Loading checkpoint: {model_spec}")
        ckpt_path = str(Path(model_spec).resolve())
        load_errors = []
        try:
            root_cls = _get_rfdetr_class("RFDETR")
            if root_cls is not None and hasattr(root_cls, "from_checkpoint"):
                return root_cls.from_checkpoint(ckpt_path)

            for cls in (RFDETRBase, RFDETRMedium, RFDETRSmall, RFDETRLarge, RFDETRNano):
                if hasattr(cls, "from_checkpoint"):
                    try:
                        return cls.from_checkpoint(ckpt_path)
                    except Exception as e:
                        load_errors.append(f"{cls.__name__}.from_checkpoint failed: {e}")

                for kw in ("pretrain_weights", "checkpoint", "weights"):
                    try:
                        return cls(**{kw: ckpt_path})
                    except TypeError:
                        continue
                    except Exception as e:
                        load_errors.append(f"{cls.__name__}({kw}=...) failed: {e}")

            details = " | ".join(load_errors[:3]) if load_errors else "No compatible loader found"
            raise RuntimeError(
                "Could not load checkpoint with installed rfdetr package. "
                f"Tried RFDETR/RFDETRBase/RFDETRMedium/RFDETRSmall/RFDETRLarge/RFDETRNano. Details: {details}"
            )
        except Exception as e:
            logger.error(f"Failed to load checkpoint: {e}")
            raise
    
    # Load by size
    models = {
        'base': RFDETRBase,
        'small': RFDETRSmall,
        'medium': RFDETRMedium,
        'large': RFDETRLarge,
        'nano': RFDETRNano,
    }
    
    cls = models.get(model_spec.lower(), RFDETRBase)
    logger.info(f"Loading {cls.__name__}")
    model = cls()
    model.eval()
    
    if torch.cuda.is_available():
        model = model.cuda()
    
    return model


def detect_frame(model, frame, conf_thresh=0.3):
    """Run inference on frame. Returns (boxes, confs)."""
    try:
        # RF-DETR wrapper APIs in this project accept file-path inference.
        with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            img_rgb = frame[:, :, ::-1]
            Image.fromarray(img_rgb).save(tmp_path, format="JPEG", quality=95)
            detections = model.predict(tmp_path, threshold=conf_thresh)
        finally:
            try:
                os.remove(tmp_path)
            except OSError:
                pass

        boxes = []
        confs = []

        if detections is None or len(detections) == 0:
            return boxes, confs

        pred_boxes = detections.xyxy.tolist()
        pred_confs = detections.confidence.tolist() if hasattr(detections, "confidence") else []

        for idx, box in enumerate(pred_boxes):
            x1, y1, x2, y2 = box[:4]
            x1 = int(max(0, x1))
            y1 = int(max(0, y1))
            x2 = int(max(0, x2))
            y2 = int(max(0, y2))
            if x2 > x1 and y2 > y1:
                boxes.append((x1, y1, x2, y2))
                confs.append(float(pred_confs[idx]) if idx < len(pred_confs) else 0.5)
        
        return boxes, confs
    
    except Exception as e:
        logger.warning(f"Detection error: {e}")
        return [], []


def main():
    parser = argparse.ArgumentParser(
        description="Standalone model evaluation on video",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Examples:\n  python model_comparison_v2.py --video video.mp4 --model base --ground-truth 25"
    )
    
    parser.add_argument('--video', required=True, help='Video file path')
    parser.add_argument('--model', required=True, help='Model (base/small/etc) or checkpoint.pth')
    parser.add_argument('--ground-truth', type=int, required=True, help='Known car count')
    parser.add_argument('--output', default='Model_Comparison.xlsx', help='Output file')
    parser.add_argument('--confidence', type=float, default=30, help='Confidence threshold (0-100, default 30%)')
    parser.add_argument('--max-distance', type=float, default=50, help='Max distance for tracker centroid matching')
    parser.add_argument('--max-age', type=int, default=30, help='Max age for unmatched tracks before removal')
    parser.add_argument(
        '--trip-wire',
        default='',
        help='Trip wire line as {x1=700;y1=300;x2=1600;y2=300}. If provided, final count uses crossings.',
    )
    
    args = parser.parse_args()

    # Normalize confidence from percentage to 0-1 range
    confidence = max(0.0, min(1.0, args.confidence / 100.0))

    # Parse optional trip wire
    trip_wire = None
    if args.trip_wire:
        try:
            trip_wire = parse_trip_wire(args.trip_wire)
            logger.info(f"Trip wire enabled: {trip_wire}")
        except ValueError as e:
            logger.error(f"Invalid --trip-wire: {e}")
            return 1
    
    # Load model
    try:
        model = load_model(args.model)
    except Exception as e:
        logger.error(f"Model loading failed: {e}")
        return 1
    
    # Open video
    video_path = Path(args.video)
    if not video_path.exists():
        logger.error(f"Video not found: {video_path}")
        return 1
    
    logger.info(f"Opening: {video_path}")
    cap = cv2.VideoCapture(str(video_path))
    if not cap.isOpened():
        logger.error("Cannot open video")
        return 1
    
    fps = cap.get(cv2.CAP_PROP_FPS) or 30
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    
    logger.info(f"Video: {total_frames} frames, {fps} FPS, {w}x{h}")
    
    # Process video
    tracker = SimpleTracker(max_distance=args.max_distance, max_age=args.max_age)
    frame_data = []
    all_confs = []
    tripwire_crossings = 0
    frame_num = 0
    
    logger.info("Processing frames...")
    start_time = time.time()
    
    while True:
        ret, frame = cap.read()
        if not ret:
            break
        
        frame_num += 1
        if frame_num % max(1, total_frames // 10) == 0:
            logger.info(f"  {frame_num}/{total_frames}")
        
        # Detect
        boxes, confs = detect_frame(model, frame, confidence)
        all_confs.extend(confs)
        
        # Track
        tracker.update(boxes)

        if trip_wire is not None:
            tripwire_crossings += tracker.count_tripwire_crossings(trip_wire[0], trip_wire[1])
        
        frame_data.append({
            'frame': frame_num,
            'boxes': len(boxes),
            'tracked': tracker.count(),
            'avg_conf': np.mean(confs) if confs else 0.0,
            'tripwire_total': tripwire_crossings,
        })
    
    cap.release()
    elapsed = time.time() - start_time
    
    # Results
    total_detected = tripwire_crossings if trip_wire is not None else tracker.count()
    ground_truth = args.ground_truth
    error = total_detected - ground_truth
    accuracy = (1.0 - abs(error) / max(ground_truth, 1)) * 100
    avg_boxes_per_frame = np.mean([f['boxes'] for f in frame_data])
    avg_conf = np.mean(all_confs) if all_confs else 0.0
    
    logger.info("\n" + "="*50)
    logger.info("RESULTS")
    logger.info("="*50)
    logger.info(f"Ground Truth:  {ground_truth}")
    logger.info(f"Detected:      {total_detected}")
    if trip_wire is not None:
        logger.info(f"Trip Crossings:{tripwire_crossings}")
    logger.info(f"Error:         {error:+d}")
    logger.info(f"Accuracy:      {accuracy:.2f}%")
    logger.info(f"Avg Conf:      {avg_conf:.4f}")
    logger.info(f"Avg Boxes/Fr:  {avg_boxes_per_frame:.2f}")
    logger.info(f"Time:          {elapsed:.1f}s")
    logger.info("="*50)
    
    # Export to Excel
    logger.info(f"Writing: {args.output}")
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Header style
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                    top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Summary section
    row = 1
    ws[f'A{row}'] = "Model Comparison Report"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    
    row = 3
    for col, title in enumerate(['Metric', 'Value'], 1):
        cell = ws.cell(row=row, column=col)
        cell.value = title
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = border
    
    metrics = [
        ('Model', args.model),
        ('Video', video_path.name),
        ('Ground Truth', ground_truth),
        ('Detected', total_detected),
        ('Trip Wire', str(trip_wire) if trip_wire is not None else 'Disabled'),
        ('Trip Crossings', tripwire_crossings if trip_wire is not None else 'N/A'),
        ('Error', error),
        ('Accuracy (%)', f'{accuracy:.2f}'),
        ('Avg Confidence', f'{avg_conf:.4f}'),
        ('Avg Boxes/Frame', f'{avg_boxes_per_frame:.2f}'),
        ('Total Frames', total_frames),
        ('FPS', f'{fps:.1f}'),
        ('Resolution', f'{w}x{h}'),
        ('Processing Time (s)', f'{elapsed:.1f}'),
    ]
    
    for label, value in metrics:
        row += 1
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].border = border
        ws[f'B{row}'].border = border
    
    # Frame details
    row += 2
    ws[f'A{row}'] = "Frame Analysis"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    
    row += 1
    for col, title in enumerate(['Frame', 'Detections', 'Tracked', 'Avg Conf', 'Tripwire Total'], 1):
        cell = ws.cell(row=row, column=col)
        cell.value = title
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.border = border
    
    for f in frame_data:
        row += 1
        ws.cell(row=row, column=1).value = f['frame']
        ws.cell(row=row, column=2).value = f['boxes']
        ws.cell(row=row, column=3).value = f['tracked']
        ws.cell(row=row, column=4).value = f'{f["avg_conf"]:.4f}'
        ws.cell(row=row, column=5).value = f.get('tripwire_total', 0)
        for col in range(1, 6):
            ws.cell(row=row, column=col).border = border
    
    # Adjust widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 18
    
    wb.save(args.output)
    logger.info(f"Done: {args.output}")
    
    return 0


if __name__ == '__main__':
    sys.exit(main())
